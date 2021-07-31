from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename='lead-list.xlsx')

ws = wb.active


class Writer:

    def __init__(self):
        self.currentRow = 2;
        self.nextRow = self.currentRow + 1;

    def insert_solar_data(self, kw, price, electric):
        ws.cell(column=10, row=self.currentRow, value=kw)
        ws.cell(column=11, row=self.currentRow, value=price)
        ws.cell(column=12, row=self.currentRow, value=electric)

        # print('\n\n ====== Inserted Solar Data {0} =======\n\n'.format(self.currentRow))

        # wb.save('dataOut.xlsx')

    def get_current_address(self):
        street = ws.cell(row=self.currentRow, column=5).value
        city = ws.cell(row=self.currentRow, column=6).value
        state = ws.cell(row=self.currentRow, column=7).value
        zip = ws.cell(row=self.currentRow, column=8).value

        return "{0} {1} {2} {3}".format(street, city, state, zip)

    def next(self):
        self.currentRow = self.currentRow + 1

        return self.currentRow

    def get_max_row(self):
        return int(ws.max_row)

    def save(self):
        wb.save('dataOut.xlsx')

        print('======== Saved Excel Data =========')
